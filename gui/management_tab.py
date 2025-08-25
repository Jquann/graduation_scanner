#!/usr/bin/env python3
"""
Student management tab for viewing and managing registered students
"""

import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import os
import sys
from PIL import Image, ImageTk
from datetime import datetime
import json
import csv
import re
import zipfile
import shutil
import tempfile
import pandas as pd
# from database import Database
from pathlib import Path

sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from config import Config


class ManagementTab:
    """Student management interface"""
    
    def __init__(self, parent, database):
        self.database = database
        
        # Create main frame
        self.frame = ttk.Frame(parent)
        
        # Sort state
        self.sort_column = None
        self.sort_reverse = False
        
        # Filter state
        self.current_filter = "all"
        self.filtered_students = []
        
        # Setup UI
        self.setup_ui()
        
        # Load initial data
        self.refresh_student_list()
    
    def setup_ui(self):
        """Setup management interface"""
        # Top section: Search and filters
        self.setup_search_section()
        
        # Middle section: Student list
        self.setup_student_list()
        
        # Bottom section: Action buttons and statistics
        self.setup_action_section()
    
    def setup_search_section(self):
        """Setup search and filter section"""
        search_frame = ttk.LabelFrame(self.frame, text="🔍 Search and Filter")
        search_frame.pack(fill='x', padx=5, pady=5)
        
        # Search bar
        search_bar_frame = ttk.Frame(search_frame)
        search_bar_frame.pack(fill='x', padx=10, pady=10)
        
        ttk.Label(search_bar_frame, text="Search:").pack(side='left', padx=5)
        
        self.search_var = tk.StringVar()
        self.search_var.trace('w', self.on_search_change)  # Auto-search on type
        self.search_entry = ttk.Entry(search_bar_frame, textvariable=self.search_var, width=30)
        self.search_entry.pack(side='left', padx=5)
        
        ttk.Button(
            search_bar_frame, 
            text="🔍 Search", 
            command=self.search_students
        ).pack(side='left', padx=5)
        
        ttk.Button(
            search_bar_frame, 
            text="🔄 Clear", 
            command=self.clear_search
        ).pack(side='left', padx=5)
        
        # Filter options
        filter_frame = ttk.Frame(search_frame)
        filter_frame.pack(fill='x', padx=10, pady=5)
        
        ttk.Label(filter_frame, text="Filter by:").pack(side='left', padx=5)
        
        self.filter_var = tk.StringVar(value="all")
        ttk.Radiobutton(
            filter_frame, 
            text="All", 
            variable=self.filter_var, 
            value="all",
            command=self.apply_filter
        ).pack(side='left', padx=5)
        
        ttk.Radiobutton(
            filter_frame, 
            text="Faculty", 
            variable=self.filter_var, 
            value="faculty",
            command=self.apply_filter
        ).pack(side='left', padx=5)
        
        ttk.Radiobutton(
            filter_frame, 
            text="Graduation Level", 
            variable=self.filter_var, 
            value="level",
            command=self.apply_filter
        ).pack(side='left', padx=5)
        
        # Faculty/Level selector (hidden by default)
        self.filter_combo_var = tk.StringVar()
        self.filter_combo = ttk.Combobox(
            filter_frame, 
            textvariable=self.filter_combo_var,
            width=25,
            state='readonly'
        )
        self.filter_combo.bind('<<ComboboxSelected>>', lambda e: self.apply_filter())
        
        # Bind Enter key to search
        self.search_entry.bind('<Return>', lambda e: self.search_students())
    
    def setup_student_list(self):
        """Setup student list tree view"""
        list_frame = ttk.LabelFrame(self.frame, text="📚 Registered Students")
        list_frame.pack(fill='both', expand=True, padx=5, pady=5)
        
        # Create tree view with scrollbars
        tree_frame = ttk.Frame(list_frame)
        tree_frame.pack(fill='both', expand=True, padx=10, pady=10)
        
        # Columns
        columns = ('ID', 'Name', 'Faculty', 'Graduation Level', 'Registration Date')
        
        self.student_tree = ttk.Treeview(
            tree_frame, 
            columns=columns, 
            show='headings',
            selectmode='browse'
        )
        
        # Configure columns
        column_widths = {
            'ID': 100,
            'Name': 200,
            'Faculty': 200,
            'Graduation Level': 200,
            'Registration Date': 150
        }
        
        for col in columns:
            self.student_tree.heading(col, text=col, command=lambda c=col: self.sort_treeview(c))
            self.student_tree.column(col, width=column_widths.get(col, 150))
        
        # Add scrollbars
        v_scrollbar = ttk.Scrollbar(tree_frame, orient="vertical", command=self.student_tree.yview)
        v_scrollbar.pack(side="right", fill="y")
        self.student_tree.configure(yscrollcommand=v_scrollbar.set)
        
        h_scrollbar = ttk.Scrollbar(tree_frame, orient="horizontal", command=self.student_tree.xview)
        h_scrollbar.pack(side="bottom", fill="x")
        self.student_tree.configure(xscrollcommand=h_scrollbar.set)
        
        self.student_tree.pack(fill='both', expand=True)
        
        # Bind double-click to view details
        self.student_tree.bind('<Double-Button-1>', lambda e: self.view_student_details())
        
        # Bind right-click for context menu
        self.student_tree.bind('<Button-3>', self.show_context_menu)
        
        # Create context menu
        self.create_context_menu()
    
    def setup_action_section(self):
        """Setup action buttons and statistics"""
        # Action buttons
        button_frame = ttk.Frame(self.frame)
        button_frame.pack(fill='x', padx=5, pady=5)
        
        ttk.Button(
            button_frame, 
            text="🔄 Refresh List", 
            command=self.refresh_student_list
        ).pack(side='left', padx=5)
        
        ttk.Button(
            button_frame, 
            text="👁️ View Details", 
            command=self.view_student_details
        ).pack(side='left', padx=5)
        
        ttk.Button(
            button_frame, 
            text="✏️ Edit Student", 
            command=self.edit_student
        ).pack(side='left', padx=5)
        
        ttk.Button(
            button_frame, 
            text="🗑️ Delete Student", 
            command=self.delete_student
        ).pack(side='left', padx=5)
        
        # Separator
        ttk.Separator(button_frame, orient='vertical').pack(side='left', fill='y', padx=10)
        
        ttk.Button(
            button_frame, 
            text="📤 Export Data", 
            command=self.export_data
        ).pack(side='left', padx=5)
        
        ttk.Button(
            button_frame, 
            text="📥 Import Data", 
            command=self.import_data
        ).pack(side='left', padx=5)
        
        # ttk.Button(
        #     button_frame, 
        #     text="💾 Backup Database", 
        #     command=self.backup_database
        # ).pack(side='left', padx=5)
        
        # Statistics frame
        stats_frame = ttk.LabelFrame(self.frame, text="📊 Statistics")
        stats_frame.pack(fill='x', padx=5, pady=(5, 0))
        
        stats_content_frame = ttk.Frame(stats_frame)
        stats_content_frame.pack(fill='x', padx=10, pady=5)
        
        # Left side statistics
        left_stats_frame = ttk.Frame(stats_content_frame)
        left_stats_frame.pack(side='left', fill='x', expand=True)
        
        self.total_students_label = ttk.Label(left_stats_frame, text="Total Students: 0", font=('Arial', 10, 'bold'))
        self.total_students_label.pack(anchor='w')
        
        self.displayed_students_label = ttk.Label(left_stats_frame, text="Displayed: 0")
        self.displayed_students_label.pack(anchor='w')
        
        # Right side statistics
        right_stats_frame = ttk.Frame(stats_content_frame)
        right_stats_frame.pack(side='right')
        
        self.faculty_stats_label = ttk.Label(right_stats_frame, text="", justify='left')
        self.faculty_stats_label.pack()
        
        self.level_stats_label = ttk.Label(right_stats_frame, text="", justify='left')
        self.level_stats_label.pack()
    
    def create_context_menu(self):
        """Create right-click context menu"""
        self.context_menu = tk.Menu(self.frame, tearoff=0)
        self.context_menu.add_command(label="👁️ View Details", command=self.view_student_details)
        self.context_menu.add_command(label="✏️ Edit Student", command=self.edit_student)
        self.context_menu.add_separator()
        self.context_menu.add_command(label="📷 View Photo", command=self.view_student_photo)
        self.context_menu.add_command(label="📱 View QR Code", command=self.view_student_qr)
        self.context_menu.add_separator()
        self.context_menu.add_command(label="🗑️ Delete Student", command=self.delete_student)
    
    def show_context_menu(self, event):
        """Show context menu on right-click"""
        # Select the item under cursor
        item = self.student_tree.identify_row(event.y)
        if item:
            self.student_tree.selection_set(item)
            self.context_menu.post(event.x_root, event.y_root)
    
    def on_search_change(self, *args):
        """Handle search text change (auto-search)"""
        # Add small delay for auto-search
        if hasattr(self, 'search_after_id'):
            self.frame.after_cancel(self.search_after_id)
        self.search_after_id = self.frame.after(300, self.search_students)
    
    def search_students(self):
        """Search students based on search criteria"""
        search_term = self.search_var.get().strip().lower()
        
        if not search_term:
            self.filtered_students = self.database.get_all_students()
        else:
            all_students = self.database.get_all_students()
            self.filtered_students = []
            
            for student in all_students:
                # Search in multiple fields
                searchable_text = f"{student.get('student_id', '')} {student.get('name', '')} {student.get('faculty', '')} {student.get('graduation_level', '')}".lower()
                
                if search_term in searchable_text:
                    self.filtered_students.append(student)
        
        self.update_student_display()
        self.update_statistics()
    
    def clear_search(self):
        """Clear search and show all students"""
        self.search_var.set("")
        self.filtered_students = self.database.get_all_students()
        self.update_student_display()
        self.update_statistics()
    
    def apply_filter(self):
        """Apply selected filter"""
        filter_type = self.filter_var.get()
        
        if filter_type == "all":
            self.filter_combo.pack_forget()
            self.filtered_students = self.database.get_all_students()
        
        elif filter_type == "faculty":
            # Show faculty selector
            self.filter_combo.pack(side='left', padx=5)
            
            # Get unique faculties
            all_students = self.database.get_all_students()
            faculties = sorted(list(set(s.get('faculty', '') for s in all_students if s.get('faculty'))))
            
            self.filter_combo['values'] = faculties
            if faculties and not self.filter_combo_var.get():
                self.filter_combo_var.set(faculties[0])
            
            # Apply faculty filter
            selected_faculty = self.filter_combo_var.get()
            if selected_faculty:
                self.filtered_students = [
                    s for s in all_students 
                    if s.get('faculty', '') == selected_faculty
                ]
            else:
                self.filtered_students = all_students
        
        elif filter_type == "level":
            # Show graduation level selector
            self.filter_combo.pack(side='left', padx=5)
            
            # Get unique graduation levels
            all_students = self.database.get_all_students()
            levels = sorted(list(set(s.get('graduation_level', '') for s in all_students if s.get('graduation_level'))))
            
            self.filter_combo['values'] = levels
            if levels and not self.filter_combo_var.get():
                self.filter_combo_var.set(levels[0])
            
            # Apply level filter
            selected_level = self.filter_combo_var.get()
            if selected_level:
                self.filtered_students = [
                    s for s in all_students 
                    if s.get('graduation_level', '') == selected_level
                ]
            else:
                self.filtered_students = all_students
        
        self.update_student_display()
        self.update_statistics()
    
    def sort_treeview(self, column):
        """Sort treeview by column"""
        if self.sort_column == column:
            self.sort_reverse = not self.sort_reverse
        else:
            self.sort_reverse = False
            self.sort_column = column
        
        # Sort the filtered students
        if column == 'Registration Date':
            self.filtered_students.sort(
                key=lambda s: s.get('registered_time', ''),
                reverse=self.sort_reverse
            )
        else:
            column_map = {
                'ID': 'student_id',
                'Name': 'name',
                'Faculty': 'faculty',
                'Graduation Level': 'graduation_level'
            }
            
            sort_key = column_map.get(column, 'student_id')
            self.filtered_students.sort(
                key=lambda s: s.get(sort_key, '').lower(),
                reverse=self.sort_reverse
            )
        
        self.update_student_display()
        
        # Update column header to show sort direction
        for col in self.student_tree['columns']:
            heading_text = col
            if col == column:
                heading_text += " ↓" if self.sort_reverse else " ↑"
            self.student_tree.heading(col, text=heading_text)
    
    def refresh_student_list(self):
        """Refresh student list from database"""
        self.filtered_students = self.database.get_all_students()
        self.update_student_display()
        self.update_statistics()
    
    def update_student_display(self):
        """Update the tree view with current filtered students"""
        # Clear existing items
        for item in self.student_tree.get_children():
            self.student_tree.delete(item)
        
        # Add filtered students
        for student in self.filtered_students:
            # Format registration date
            reg_date = student.get('registered_time', '')
            if reg_date:
                try:
                    dt = datetime.fromisoformat(reg_date.replace('Z', '+00:00'))
                    formatted_date = dt.strftime('%Y-%m-%d %H:%M')
                except:
                    formatted_date = reg_date[:16] if len(reg_date) > 16 else reg_date
            else:
                formatted_date = 'Unknown'
            
            self.student_tree.insert("", "end", values=(
                student.get("student_id", ""),
                student.get("name", ""),
                student.get("faculty", ""),
                student.get("graduation_level", ""),
                formatted_date
            ))
    
    def update_statistics(self):
        """Update statistics display"""
        total_students = len(self.database.get_all_students())
        displayed_students = len(self.filtered_students)
        
        self.total_students_label.config(text=f"Total Students: {total_students}")
        self.displayed_students_label.config(text=f"Displayed: {displayed_students}")
        
        # Faculty statistics
        if self.filtered_students:
            faculty_counts = {}
            level_counts = {}
            
            for student in self.filtered_students:
                faculty = student.get('faculty', 'Unknown')
                level = student.get('graduation_level', 'Unknown')
                
                faculty_counts[faculty] = faculty_counts.get(faculty, 0) + 1
                level_counts[level] = level_counts.get(level, 0) + 1
            
            # Format faculty stats
            faculty_text = "Faculties:\n" + "\n".join(
                f"• {faculty}: {count}" 
                for faculty, count in sorted(faculty_counts.items())
            )
            
            # Format level stats
            level_text = "Graduation Levels:\n" + "\n".join(
                f"• {level}: {count}" 
                for level, count in sorted(level_counts.items())
            )
            
            self.faculty_stats_label.config(text=faculty_text)
            self.level_stats_label.config(text=level_text)
        else:
            self.faculty_stats_label.config(text="")
            self.level_stats_label.config(text="")
    
    def get_selected_student(self):
        """Get currently selected student data"""
        selected_item = self.student_tree.selection()
        if not selected_item:
            return None
        
        item_values = self.student_tree.item(selected_item[0])['values']
        student_id = item_values[0]

        # Find student in filtered list
        for student in self.filtered_students:
            if student.get('student_id') == student_id:
                return student
        
        # Fallback: search in entire database
        return self.database.get_student_by_id(student_id)
    
    def view_student_details(self):
        """View detailed student information"""
        student = self.get_selected_student()
        if not student:
            messagebox.showwarning("Warning", "Please select a student to view details")
            return
        
        self.show_student_details_window(student)
    
    def show_student_details_window(self, student):
        """Show student details in a new window"""
        detail_window = tk.Toplevel(self.frame)
        detail_window.title(f"Student Details - {student.get('name', 'Unknown')}")
        detail_window.geometry("800x700")
        detail_window.resizable(True, True)
        
        # Create notebook for tabs
        notebook = ttk.Notebook(detail_window)
        notebook.pack(fill='both', expand=True, padx=10, pady=10)
        
        # Basic Information Tab
        self.create_basic_info_tab(notebook, student)
        
        # Media Tab (Photo & QR)
        self.create_media_tab(notebook, student)
        
        # Technical Tab
        self.create_technical_tab(notebook, student)
    
    def create_basic_info_tab(self, notebook, student):
        """Create basic information tab"""
        basic_frame = ttk.Frame(notebook)
        notebook.add(basic_frame, text="📋 Basic Info")
        
        # Student Information
        info_frame = ttk.LabelFrame(basic_frame, text="Student Information")
        info_frame.pack(fill='x', padx=10, pady=10)
        
        info_items = [
            ("Student ID:", student.get("student_id", "N/A")),
            ("Full Name:", student.get("name", "N/A")),
            ("Faculty/Department:", student.get("faculty", "N/A")),
            ("Graduation Level:", student.get("graduation_level", "N/A")),
            ("Registration Date:", self.format_datetime(student.get("registered_time", ""))),
        ]
        
        for i, (label, value) in enumerate(info_items):
            ttk.Label(info_frame, text=label, font=('Arial', 10, 'bold')).grid(
                row=i, column=0, sticky='w', padx=10, pady=5
            )
            ttk.Label(info_frame, text=str(value), wraplength=300).grid(
                row=i, column=1, sticky='w', padx=20, pady=5
            )
        
        # File Paths
        paths_frame = ttk.LabelFrame(basic_frame, text="File Paths")
        paths_frame.pack(fill='x', padx=10, pady=10)
        
        path_items = [
            ("Photo Path:", student.get("photo_path", "N/A")),
            ("QR Code Path:", student.get("qr_code_path", "N/A")),
        ]
        
        for i, (label, path) in enumerate(path_items):
            ttk.Label(paths_frame, text=label, font=('Arial', 10, 'bold')).grid(
                row=i, column=0, sticky='w', padx=10, pady=5
            )
            
            path_frame = ttk.Frame(paths_frame)
            path_frame.grid(row=i, column=1, sticky='ew', padx=20, pady=5)
            
            ttk.Label(path_frame, text=str(path), wraplength=400).pack(side='left')
            
            if path and path != "N/A" and os.path.exists(path):
                ttk.Button(
                    path_frame, 
                    text="📁 Open", 
                    command=lambda p=path: self.open_file_location(p)
                ).pack(side='right', padx=5)
    
    def create_media_tab(self, notebook, student):
        """Create media tab with photo and QR code"""
        media_frame = ttk.Frame(notebook)
        notebook.add(media_frame, text="📷 Media")
        
        # Create two columns
        left_frame = ttk.LabelFrame(media_frame, text="📸 Student Photo")
        left_frame.pack(side='left', fill='both', expand=True, padx=5, pady=10)
        
        right_frame = ttk.LabelFrame(media_frame, text="📱 QR Code")
        right_frame.pack(side='right', fill='both', expand=True, padx=5, pady=10)
        
        # Load and display photo
        photo_path = student.get("photo_path", "")
        if photo_path and os.path.exists(photo_path):
            try:
                photo = Image.open(photo_path)
                photo = photo.resize((300, 300), Image.Resampling.LANCZOS)
                photo_tk = ImageTk.PhotoImage(photo)
                
                photo_label = ttk.Label(left_frame, image=photo_tk)
                photo_label.image = photo_tk
                photo_label.pack(pady=10)
                
                # Photo info
                photo_info = f"Size: {photo.size[0]}x{photo.size[1]}\nFile: {os.path.basename(photo_path)}"
                ttk.Label(left_frame, text=photo_info, justify='center').pack()
                
            except Exception as e:
                ttk.Label(left_frame, text=f"Cannot load photo:\n{str(e)}").pack(pady=50)
        else:
            ttk.Label(left_frame, text="No photo available").pack(pady=50)
        
        # Load and display QR code
        qr_path = student.get("qr_code_path", "")
        if qr_path and os.path.exists(qr_path):
            try:
                qr_img = Image.open(qr_path)
                qr_img = qr_img.resize((300, 300), Image.Resampling.NEAREST)
                qr_tk = ImageTk.PhotoImage(qr_img)
                
                qr_label = ttk.Label(right_frame, image=qr_tk)
                qr_label.image = qr_tk
                qr_label.pack(pady=10)
                
                # QR info
                qr_info = f"Student ID: {student.get('student_id', 'N/A')}\nFile: {os.path.basename(qr_path)}"
                ttk.Label(right_frame, text=qr_info, justify='center').pack()
                
                # Print button
                ttk.Button(
                    right_frame, 
                    text="🖨️ Save for Printing", 
                    command=lambda: self.save_qr_for_printing(student)
                ).pack(pady=5)
                
            except Exception as e:
                ttk.Label(right_frame, text=f"Cannot load QR code:\n{str(e)}").pack(pady=50)
        else:
            ttk.Label(right_frame, text="No QR code available").pack(pady=50)
    
    def create_technical_tab(self, notebook, student):
        """Create technical information tab"""
        tech_frame = ttk.Frame(notebook)
        notebook.add(tech_frame, text="🔧 Technical")
        
        # Face encoding information
        encoding_frame = ttk.LabelFrame(tech_frame, text="Face Encoding Information")
        encoding_frame.pack(fill='x', padx=10, pady=10)
        
        face_encoding = student.get("face_encoding", [])
        if face_encoding:
            encoding_info = [
                ("Encoding Length:", len(face_encoding)),
                ("Encoding Type:", "InsightFace buffalo_l"),
                ("Vector Dimensions:", "512D normalized embedding"),
                ("First 5 values:", str([f"{x:.6f}" for x in face_encoding[:5]])),
                ("Last 5 values:", str([f"{x:.6f}" for x in face_encoding[-5:]])),
            ]
        else:
            encoding_info = [("Status:", "No face encoding available")]
        
        for i, (label, value) in enumerate(encoding_info):
            ttk.Label(encoding_frame, text=label, font=('Arial', 10, 'bold')).grid(
                row=i, column=0, sticky='nw', padx=10, pady=5
            )
            ttk.Label(encoding_frame, text=str(value), wraplength=400).grid(
                row=i, column=1, sticky='w', padx=20, pady=5
            )
        
        # Database information
        db_frame = ttk.LabelFrame(tech_frame, text="Database Information")
        db_frame.pack(fill='x', padx=10, pady=10)
        
        # Convert datetime for display
        reg_time = student.get("registered_time", "")
        formatted_reg_time = self.format_datetime(reg_time)
        
        db_info = [
            ("Record Created:", formatted_reg_time),
            ("Database Version:", "JSON Format v1.0"),
            ("Backup Status:", "✅ Available" if os.path.exists(self.database.data_file) else "❌ Not Available"),
        ]
        
        for i, (label, value) in enumerate(db_info):
            ttk.Label(db_frame, text=label, font=('Arial', 10, 'bold')).grid(
                row=i, column=0, sticky='w', padx=10, pady=5
            )
            ttk.Label(db_frame, text=str(value)).grid(
                row=i, column=1, sticky='w', padx=20, pady=5
            )
    
    def format_datetime(self, datetime_str):
        """Format datetime string for display"""
        if not datetime_str:
            return "Unknown"
        
        try:
            dt = datetime.fromisoformat(datetime_str.replace('Z', '+00:00'))
            return dt.strftime('%Y-%m-%d %H:%M:%S')
        except:
            return datetime_str
    
    def open_file_location(self, file_path):
        """Open file location in system file manager"""
        try:
            import subprocess
            import platform
            
            if platform.system() == "Windows":
                subprocess.run(["explorer", "/select,", file_path])
            elif platform.system() == "Darwin":  # macOS
                subprocess.run(["open", "-R", file_path])
            else:  # Linux
                subprocess.run(["xdg-open", os.path.dirname(file_path)])
        except Exception as e:
            messagebox.showerror("Error", f"Cannot open file location: {e}")

    def save_qr_for_printing(self, student):
        """Save QR code in print-ready format"""
        qr_path = student.get("qr_code_path", "")
        if not qr_path or not os.path.exists(qr_path):
            messagebox.showerror("Error", "QR code file not found")
            return
        
        # Create default filename using student info
        student_id = student.get('student_id', 'Unknown')
        student_name = student.get('name', 'Unknown')
        
        # Clean the name for filename (remove special characters)
        clean_name = re.sub(r'[<>:"/\\|?*]', '', student_name)  # Remove invalid filename characters
        clean_name = clean_name.replace(' ', '_')  # Replace spaces with underscores
        
        # Create default filename
        default_filename = f"QR_Code_{student_id}_{clean_name}.png"
        
        # Ask user for save location with default filename
        save_path = filedialog.asksaveasfilename(
            title=f"Save QR Code for {student.get('name', 'Unknown')}",
            initialfile=default_filename,  # This sets the default filename
            defaultextension=".png",
            filetypes=[("PNG files", "*.png"), ("All files", "*.*")]
        )
        
        if save_path:
            try:
                # Create a print-ready version with student info
                qr_img = Image.open(qr_path)
                
                # Create a larger canvas with student info
                from PIL import ImageDraw, ImageFont
                
                canvas_width = 400
                canvas_height = 500
                canvas = Image.new('RGB', (canvas_width, canvas_height), 'white')
                
                # Resize QR code
                qr_size = 300
                qr_img = qr_img.resize((qr_size, qr_size), Image.Resampling.NEAREST)
                
                # Paste QR code onto canvas
                qr_x = (canvas_width - qr_size) // 2
                qr_y = 50
                canvas.paste(qr_img, (qr_x, qr_y))
                
                # Add text information
                draw = ImageDraw.Draw(canvas)
                
                try:
                    font_large = ImageFont.truetype("arial.ttf", 16)
                    font_small = ImageFont.truetype("arial.ttf", 12)
                except:
                    font_large = ImageFont.load_default()
                    font_small = ImageFont.load_default()
                
                # Student information
                text_y = qr_y + qr_size + 20
                student_info = [
                    f"Student ID: {student.get('student_id', 'N/A')}",
                    f"Name: {student.get('name', 'N/A')}",
                    f"Faculty: {student.get('faculty', 'N/A')}"
                ]
                
                for info in student_info:
                    text_bbox = draw.textbbox((0, 0), info, font=font_small)
                    text_width = text_bbox[2] - text_bbox[0]
                    text_x = (canvas_width - text_width) // 2
                    draw.text((text_x, text_y), info, fill='black', font=font_small)
                    text_y += 25
                
                # Save the canvas
                canvas.save(save_path, 'PNG', dpi=(300, 300))
                messagebox.showinfo("Success", f"QR code saved for printing:\n{save_path}")
                
            except Exception as e:
                messagebox.showerror("Error", f"Error saving QR code: {e}")
                
    def edit_student(self):
        """Edit selected student information"""
        student = self.get_selected_student()
        if not student:
            messagebox.showwarning("Warning", "Please select a student to edit")
            return
        
        self.show_edit_student_window(student)
    
    def show_edit_student_window(self, student):
        """Show edit student window"""
        edit_window = tk.Toplevel(self.frame)
        edit_window.title(f"Edit Student - {student.get('name', 'Unknown')}")
        edit_window.geometry("600x500")
        edit_window.resizable(False, False)
        
        # Make window modal
        edit_window.transient(self.frame)
        edit_window.grab_set()
        
        # Create form
        form_frame = ttk.LabelFrame(edit_window, text="Student Information")
        form_frame.pack(fill='both', expand=True, padx=10, pady=10)
        
        # Student ID (read-only)
        ttk.Label(form_frame, text="Student ID:").grid(row=0, column=0, sticky='w', padx=10, pady=5)
        id_label = ttk.Label(form_frame, text=student.get("student_id", ""), font=('Arial', 10, 'bold'))
        id_label.grid(row=0, column=1, sticky='w', padx=10, pady=5)
        ttk.Label(form_frame, text="(Cannot be changed)", font=('Arial', 8), foreground='gray').grid(row=0, column=2, sticky='w', padx=5, pady=5)
        
        # Name
        ttk.Label(form_frame, text="Full Name:").grid(row=1, column=0, sticky='w', padx=10, pady=5)
        name_var = tk.StringVar(value=student.get("name", ""))
        name_entry = ttk.Entry(form_frame, textvariable=name_var, width=40)
        name_entry.grid(row=1, column=1, columnspan=2, sticky='w', padx=10, pady=5)
        
        # Faculty
        ttk.Label(form_frame, text="Faculty:").grid(row=2, column=0, sticky='w', padx=10, pady=5)
        faculty_var = tk.StringVar(value=student.get("faculty", ""))
        faculty_entry = ttk.Entry(form_frame, textvariable=faculty_var, width=40)
        faculty_entry.grid(row=2, column=1, columnspan=2, sticky='w', padx=10, pady=5)
        
        # Graduation Level
        ttk.Label(form_frame, text="Graduation Level:").grid(row=3, column=0, sticky='w', padx=10, pady=5)
        level_var = tk.StringVar(value=student.get("graduation_level", ""))
        level_combo = ttk.Combobox(form_frame, textvariable=level_var, width=37)
        level_combo['values'] = ('Pass', 'With Merit', 'With Distinction', 'With Distinction and Book Prize')
        level_combo.grid(row=3, column=1, columnspan=2, sticky='w', padx=10, pady=5)
        
        # Current photo display
        ttk.Label(form_frame, text="Current Photo:").grid(row=4, column=0, sticky='nw', padx=10, pady=5)
        
        photo_frame = ttk.Frame(form_frame)
        photo_frame.grid(row=4, column=1, columnspan=2, sticky='w', padx=10, pady=5)
        
        # Display current photo
        photo_path = student.get("photo_path", "")
        if photo_path and os.path.exists(photo_path):
            try:
                photo = Image.open(photo_path)
                photo = photo.resize((100, 100), Image.Resampling.LANCZOS)
                photo_tk = ImageTk.PhotoImage(photo)
                
                photo_label = ttk.Label(photo_frame, image=photo_tk)
                photo_label.image = photo_tk
                photo_label.pack(side='left')
            except:
                ttk.Label(photo_frame, text="Cannot load photo").pack(side='left')
        else:
            ttk.Label(photo_frame, text="No photo").pack(side='left')
        
        # Photo change button
        ttk.Button(photo_frame, text="📷 Change Photo", 
                  command=lambda: self.change_student_photo(student, photo_frame)).pack(side='left', padx=10)
        
        # Buttons
        button_frame = ttk.Frame(edit_window)
        button_frame.pack(fill='x', padx=10, pady=10)
        
        def save_changes():
            # Validate input
            new_name = name_var.get().strip()
            new_faculty = faculty_var.get().strip()
            new_level = level_var.get().strip()
            
            if not all([new_name, new_faculty, new_level]):
                messagebox.showerror("Error", "Please fill in all required fields")
                return
            
            # Update student data
            updated_student = student.copy()
            updated_student["name"] = new_name
            updated_student["faculty"] = new_faculty
            updated_student["graduation_level"] = new_level
            
            # Save to database
            try:
                if hasattr(self.database, 'update_student'):
                    self.database.update_student(student["student_id"], updated_student)
                else:
                    # For simple JSON DB without update method
                    students = self.database.get_all_students()
                    for i, s in enumerate(students):
                        if s.get('student_id') == student["student_id"]:
                            students[i] = updated_student
                            break
                    self.database.save_students(students)

                self.refresh_student_list()
                edit_window.destroy()
                messagebox.showinfo("Success", "Student information updated successfully!")
            except Exception as e:
                messagebox.showerror("Error", f"Error updating student: {e}")
        
        ttk.Button(button_frame, text="💾 Save Changes", command=save_changes).pack(side='right', padx=5)
        ttk.Button(button_frame, text="❌ Cancel", command=edit_window.destroy).pack(side='right', padx=5)
        
        # Focus on name field
        name_entry.focus_set()
    
    def change_student_photo(self, student, photo_frame):
        """Change student photo"""
        new_photo_path = filedialog.askopenfilename(
            title="Select New Photo",
            filetypes=[("Image files", "*.jpg *.jpeg *.png *.bmp")]
        )
        
        if new_photo_path:
            try:
                # Process the new photo (similar to registration)
                import cv2
                from face_recognition import FaceRecognition
                
                # Initialize face recognition
                face_app = FaceRecognition()
                
                # Read and process image
                image = cv2.imread(new_photo_path)
                image_rgb = cv2.cvtColor(image, cv2.COLOR_BGR2RGB)
                
                faces = face_app.get(image_rgb)
                if not faces:
                    messagebox.showerror("Error", "No face detected in the new photo")
                    return
                
                if len(faces) > 1:
                    messagebox.showwarning("Warning", "Multiple faces detected, using the first one")
                
                # Get face encoding
                face_encoding = faces[0].normed_embedding.tolist()
                
                # Save new photo
                from pathlib import Path
                photos_dir = Path("graduation_data/photos")
                photos_dir.mkdir(exist_ok=True)
                
                photo_filename = f"{student['student_id']}.jpg"
                photo_save_path = photos_dir / photo_filename
                cv2.imwrite(str(photo_save_path), image)
                
                # Update student data
                updated_student = student.copy()
                updated_student["photo_path"] = str(photo_save_path)
                updated_student["face_encoding"] = face_encoding
                
                self.database.update_student(student["student_id"], updated_student)
                
                # Update photo display
                for widget in photo_frame.winfo_children():
                    if isinstance(widget, ttk.Label) and hasattr(widget, 'image'):
                        widget.destroy()
                        break
                
                # Display new photo
                photo = Image.open(photo_save_path)
                photo = photo.resize((100, 100), Image.Resampling.LANCZOS)
                photo_tk = ImageTk.PhotoImage(photo)
                
                new_photo_label = ttk.Label(photo_frame, image=photo_tk)
                new_photo_label.image = photo_tk
                new_photo_label.pack(side='left')
                
                messagebox.showinfo("Success", "Photo updated successfully!")
                
            except Exception as e:
                messagebox.showerror("Error", f"Error updating photo: {e}")
    
    def delete_student(self):
        """Delete selected student"""
        student = self.get_selected_student()
        if not student:
            messagebox.showwarning("Warning", "Please select a student to delete")
            return
        
        student_name = student.get("name", "Unknown")
        student_id = student.get("student_id", "Unknown")
        
        if messagebox.askyesno(
            "Confirm Delete", 
            f"Are you sure you want to delete student:\n\n"
            f"ID: {student_id}\n"
            f"Name: {student_name}\n\n"
            f"This action cannot be undone!"
        ):
            try:
                self.database.delete_student(student_id)
                self.refresh_student_list()
                messagebox.showinfo("Success", "Student deleted successfully")
            except Exception as e:
                messagebox.showerror("Error", f"Error deleting student: {e}")
    
    def view_student_photo(self):
        """View student photo in a separate window"""
        student = self.get_selected_student()
        if not student:
            messagebox.showwarning("Warning", "Please select a student")
            return
        
        photo_path = student.get("photo_path", "")
        if not photo_path or not os.path.exists(photo_path):
            messagebox.showwarning("Warning", "No photo available for this student")
            return
        
        # Create photo viewer window
        photo_window = tk.Toplevel(self.frame)
        photo_window.title(f"Photo - {student.get('name', 'Unknown')}")
        photo_window.geometry("500x600")
        
        try:
            photo = Image.open(photo_path)
            # Scale to fit window while maintaining aspect ratio
            photo.thumbnail((450, 450), Image.Resampling.LANCZOS)
            photo_tk = ImageTk.PhotoImage(photo)
            
            photo_label = ttk.Label(photo_window, image=photo_tk)
            photo_label.image = photo_tk
            photo_label.pack(pady=20)
            
            # Photo info
            info_text = f"Student: {student.get('name', 'Unknown')}\nID: {student.get('student_id', 'Unknown')}\nFile: {os.path.basename(photo_path)}"
            ttk.Label(photo_window, text=info_text, justify='center').pack(pady=10)
            
        except Exception as e:
            ttk.Label(photo_window, text=f"Error loading photo: {e}").pack(pady=50)
    
    def view_student_qr(self):
        """View student QR code in a separate window"""
        student = self.get_selected_student()
        if not student:
            messagebox.showwarning("Warning", "Please select a student")
            return
        
        qr_path = student.get("qr_code_path", "")
        if not qr_path or not os.path.exists(qr_path):
            messagebox.showwarning("Warning", "No QR code available for this student")
            return
        
        # Create QR viewer window
        qr_window = tk.Toplevel(self.frame)
        qr_window.title(f"QR Code - {student.get('name', 'Unknown')}")
        qr_window.geometry("400x500")
        
        try:
            qr_img = Image.open(qr_path)
            qr_img = qr_img.resize((300, 300), Image.Resampling.NEAREST)
            qr_tk = ImageTk.PhotoImage(qr_img)
            
            qr_label = ttk.Label(qr_window, image=qr_tk)
            qr_label.image = qr_tk
            qr_label.pack(pady=20)
            
            # QR info
            info_text = f"Student: {student.get('name', 'Unknown')}\nID: {student.get('student_id', 'Unknown')}\nQR Data: {student.get('student_id', 'Unknown')}"
            ttk.Label(qr_window, text=info_text, justify='center').pack(pady=10)
            
            # Save button
            ttk.Button(qr_window, text="💾 Save QR Code", 
                      command=lambda: self.save_qr_for_printing(student)).pack(pady=10)
            
        except Exception as e:
            ttk.Label(qr_window, text=f"Error loading QR code: {e}").pack(pady=50)

    def export_data(self):
        """Export student data with graduation_data folder"""
        if not self.filtered_students:
            messagebox.showwarning("Warning", "No student data to export")
            return
        
        # Create export options window
        export_window = tk.Toplevel(self.frame)
        export_window.title("Export Student Data")
        export_window.geometry("450x400")
        export_window.resizable(False, False)
        
        # Make modal
        export_window.transient(self.frame)
        export_window.grab_set()
        
        # Export format selection
        format_frame = ttk.LabelFrame(export_window, text="Export Format")
        format_frame.pack(fill='x', padx=10, pady=10)
        
        format_var = tk.StringVar(value="zip_full")
        
        # ttk.Radiobutton(format_frame, text="Complete Backup (ZIP with all files)", 
        #             variable=format_var, value="zip_full").pack(anchor='w', padx=10, pady=5)
        ttk.Radiobutton(format_frame, text="Data Only - CSV", 
                    variable=format_var, value="csv").pack(anchor='w', padx=10, pady=5)
        ttk.Radiobutton(format_frame, text="Data Only - JSON", 
                    variable=format_var, value="json").pack(anchor='w', padx=10, pady=5)
        ttk.Radiobutton(format_frame, text="Data Only - Excel (.xlsx)", 
                    variable=format_var, value="xlsx").pack(anchor='w', padx=10, pady=5)
        
        # Data selection for non-ZIP exports
        data_frame = ttk.LabelFrame(export_window, text="Data Options (for non-ZIP exports)")
        data_frame.pack(fill='x', padx=10, pady=10)
        
        include_paths_var = tk.BooleanVar(value=False)
        include_technical_var = tk.BooleanVar(value=False)
        
        ttk.Checkbutton(data_frame, text="Include file paths", 
                    variable=include_paths_var).pack(anchor='w', padx=10, pady=5)
        ttk.Checkbutton(data_frame, text="Include technical data (face encodings)", 
                    variable=include_technical_var).pack(anchor='w', padx=10, pady=5)
        
        # Export info
        info_frame = ttk.Frame(export_window)
        info_frame.pack(fill='x', padx=10, pady=10)
        
        ttk.Label(info_frame, text=f"Students to export: {len(self.filtered_students)}", 
                font=('Arial', 10, 'bold')).pack()
        
        # Description label
        # desc_text = "ZIP export includes all photos, QR codes, and database files"
        # ttk.Label(info_frame, text=desc_text, 
        #         font=('Arial', 9), foreground='blue').pack(pady=5)
        
        # Buttons
        button_frame = ttk.Frame(export_window)
        button_frame.pack(fill='x', padx=10, pady=10)
        
        def perform_export():
            export_format = format_var.get()
            include_paths = include_paths_var.get()
            include_technical = include_technical_var.get()
            
            if export_format == "zip_full":
                # ZIP export
                save_path = filedialog.asksaveasfilename(
                    title="Export Complete Backup",
                    filetypes=[("ZIP files", "*.zip"), ("All files", "*.*")],
                    defaultextension=".zip"
                )
            else:
                # Data-only export
                file_types = {
                    "csv": [("CSV files", "*.csv")],
                    "json": [("JSON files", "*.json")],
                    "xlsx": [("Excel files", "*.xlsx")]
                }
                
                save_path = filedialog.asksaveasfilename(
                    title="Export Student Data",
                    filetypes=file_types[export_format] + [("All files", "*.*")],
                    defaultextension=f".{export_format}"
                )
            
            if save_path:
                try:
                    if export_format == "zip_full":
                        self.export_complete_backup(save_path)
                    else:
                        self.perform_data_export(save_path, export_format, include_paths, include_technical)
                    
                    export_window.destroy()
                    messagebox.showinfo("Success", f"Data exported successfully to:\n{save_path}")
                except Exception as e:
                    messagebox.showerror("Error", f"Export failed: {e}")
        
        ttk.Button(button_frame, text="📤 Export", command=perform_export).pack(side='right', padx=5)
        ttk.Button(button_frame, text="❌ Cancel", command=export_window.destroy).pack(side='right', padx=5)

    
    def export_complete_backup(self, zip_path):
        """Export complete backup including graduation_data folder - FIXED VERSION"""
        graduation_data_path = Path("graduation_data")
        
        if not graduation_data_path.exists():
            raise FileNotFoundError("graduation_data folder not found")
        
        # Create progress window
        progress_window = tk.Toplevel(self.frame)
        progress_window.title("Exporting...")
        progress_window.geometry("400x150")
        progress_window.resizable(False, False)
        progress_window.transient(self.frame)
        progress_window.grab_set()
        
        progress_label = ttk.Label(progress_window, text="Preparing export...")
        progress_label.pack(pady=20)
        
        progress_var = tk.DoubleVar()
        progress_bar = ttk.Progressbar(progress_window, variable=progress_var, maximum=100)
        progress_bar.pack(pady=10, padx=20, fill='x')
        
        progress_window.update()
        
        try:
            with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as zipf:
                # Get all files to zip
                all_files = []
                for root, dirs, files in os.walk(graduation_data_path):
                    for file in files:
                        file_path = Path(root) / file
                        all_files.append(file_path)
                
                total_files = len(all_files)
                
                if total_files == 0:
                    raise ValueError("No files found in graduation_data folder")
                
                processed_files = 0
                
                for file_path in all_files:
                    # Update progress
                    progress_label.config(text=f"Adding: {file_path.name}")
                    progress_var.set((processed_files / total_files) * 100)
                    progress_window.update()
                    
                    # Add file to zip with relative path from current directory
                    # This ensures the zip contains "graduation_data/..." structure
                    relative_path = file_path.relative_to(Path.cwd())
                    zipf.write(file_path, relative_path)
                    processed_files += 1
                
                # Add export metadata
                metadata = {
                    "export_timestamp": datetime.now().isoformat(),
                    "total_students": len(self.database.get_all_students()),
                    "export_type": "complete_backup",
                    "version": "1.0",
                    "application": "graduation_scanner"
                }
                
                zipf.writestr("export_metadata.json", json.dumps(metadata, indent=2))
                
            progress_label.config(text="Export completed!")
            progress_var.set(100)
            progress_window.update()
            
            # Close progress window after a short delay
            progress_window.after(1000, progress_window.destroy)
            
        except Exception as e:
            progress_window.destroy()
            raise e


    def perform_data_export(self, file_path, format_type, include_paths, include_technical):
        """Perform the actual data export"""
        # Prepare data
        export_data = []
        
        for student in self.filtered_students:
            student_data = {
                "student_id": student.get("student_id", ""),
                "name": student.get("name", ""),
                "faculty": student.get("faculty", ""),
                "graduation_level": student.get("graduation_level", ""),
                "registration_date": self.format_datetime(student.get("registered_time", ""))
            }
            
            if include_paths:
                student_data.update({
                    "photo_path": student.get("photo_path", ""),
                    "qr_code_path": student.get("qr_code_path", "")
                })
            
            if include_technical:
                student_data.update({
                    "face_encoding": student.get("face_encoding", []),
                    "registered_time": student.get("registered_time", "")
                })
            
            export_data.append(student_data)
        
        # Export based on format
        if format_type == "csv":
            self.export_to_csv(file_path, export_data)
        elif format_type == "json":
            self.export_to_json(file_path, export_data)
        elif format_type == "xlsx":
            self.export_to_xlsx(file_path, export_data)
    
    def export_to_csv(self, file_path, data):
        """Export data to CSV format"""
        if not data:
            return
        
        with open(file_path, 'w', newline='', encoding='utf-8') as csvfile:
            fieldnames = data[0].keys()
            writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
            
            writer.writeheader()
            for row in data:
                # Handle face_encoding for CSV (convert to string)
                if 'face_encoding' in row and isinstance(row['face_encoding'], list):
                    row['face_encoding'] = str(row['face_encoding'])
                writer.writerow(row)
    
    def export_to_json(self, file_path, data):
        """Export data to JSON format"""
        export_structure = {
            "export_info": {
                "timestamp": datetime.now().isoformat(),
                "total_students": len(data),
                "export_type": "graduation_scanner_export"
            },
            "students": data
        }
        
        with open(file_path, 'w', encoding='utf-8') as jsonfile:
            json.dump(export_structure, jsonfile, indent=2, ensure_ascii=False)
    
    def export_to_xlsx(self, file_path, data):
        """Export data to Excel format"""
        try:
            import openpyxl
            from openpyxl.utils import get_column_letter
            from openpyxl.styles import Font, PatternFill, Alignment
            
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Student Data"
            
            if not data:
                wb.save(file_path)
                return
            
            # Headers
            headers = list(data[0].keys())
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header.replace('_', ' ').title())
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")
            
            # Data
            for row_idx, student in enumerate(data, 2):
                for col_idx, (key, value) in enumerate(student.items(), 1):
                    if key == 'face_encoding' and isinstance(value, list):
                        value = f"[{len(value)} values]"
                    ws.cell(row=row_idx, column=col_idx, value=value)
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            wb.save(file_path)
            
        except ImportError:
            messagebox.showerror("Error", "openpyxl library not installed. Cannot export to Excel format.")
            raise
  
    def import_data(self):
        """Import student data from file"""
        # Create import options window
        import_window = tk.Toplevel(self.frame)
        import_window.title("Import Student Data")
        import_window.geometry("500x480")
        import_window.resizable(False, False)
        
        # Make modal
        import_window.transient(self.frame)
        import_window.grab_set()
        
        # Import format selection
        format_frame = ttk.LabelFrame(import_window, text="Import Format")
        format_frame.pack(fill='x', padx=10, pady=10)
        
        format_var = tk.StringVar(value="zip")
        
        # ttk.Radiobutton(format_frame, text="Complete Backup (ZIP file)", 
        #             variable=format_var, value="zip").pack(anchor='w', padx=10, pady=5)
        ttk.Radiobutton(format_frame, text="CSV Data File", 
                    variable=format_var, value="csv").pack(anchor='w', padx=10, pady=5)
        ttk.Radiobutton(format_frame, text="JSON Data File", 
                    variable=format_var, value="json").pack(anchor='w', padx=10, pady=5)
        
        # Import options
        options_frame = ttk.LabelFrame(import_window, text="Import Options")
        options_frame.pack(fill='x', padx=10, pady=10)
        
        merge_var = tk.StringVar(value="merge")
        
        ttk.Radiobutton(options_frame, text="Merge with existing data (keep both)", 
                    variable=merge_var, value="merge").pack(anchor='w', padx=10, pady=5)
        ttk.Radiobutton(options_frame, text="Replace existing data (backup current first)", 
                    variable=merge_var, value="replace").pack(anchor='w', padx=10, pady=5)
        
        # Import validation options
        validation_frame = ttk.LabelFrame(import_window, text="Validation Options")
        validation_frame.pack(fill='x', padx=10, pady=10)
        
        skip_duplicates_var = tk.BooleanVar(value=True)
        validate_data_var = tk.BooleanVar(value=True)
        
        ttk.Checkbutton(validation_frame, text="Skip duplicate student IDs", 
                    variable=skip_duplicates_var).pack(anchor='w', padx=10, pady=5)
        ttk.Checkbutton(validation_frame, text="Validate data format", 
                    variable=validate_data_var).pack(anchor='w', padx=10, pady=5)
        
        # File selection
        file_frame = ttk.Frame(import_window)
        file_frame.pack(fill='x', padx=10, pady=10)
        
        file_path_var = tk.StringVar()
        ttk.Label(file_frame, text="Selected file:").pack(anchor='w')
        file_entry = ttk.Entry(file_frame, textvariable=file_path_var, width=50)
        file_entry.pack(side='left', fill='x', expand=True, pady=5)
        
        def browse_file():
            format_type = format_var.get()
            file_types = {
                "zip": [("ZIP files", "*.zip")],
                "csv": [("CSV files", "*.csv")],
                "json": [("JSON files", "*.json")]
            }
            
            file_path = filedialog.askopenfilename(
                title=f"Select {format_type.upper()} file to import",
                filetypes=file_types[format_type] + [("All files", "*.*")]
            )
            
            if file_path:
                file_path_var.set(file_path)
        
        ttk.Button(file_frame, text="📁 Browse", command=browse_file).pack(side='right', padx=5)
        
        # Warning label
        warning_frame = ttk.Frame(import_window)
        warning_frame.pack(fill='x', padx=10, pady=5)
        
        warning_text = "⚠️ Warning: Import will modify your database. Consider backing up first."
        ttk.Label(warning_frame, text=warning_text, 
                font=('Arial', 9), foreground='red').pack()
        
        # Buttons
        button_frame = ttk.Frame(import_window)
        button_frame.pack(fill='x', padx=10, pady=10)
        
        def perform_import():
            file_path = file_path_var.get().strip()
            if not file_path:
                messagebox.showerror("Error", "Please select a file to import")
                return
            
            if not os.path.exists(file_path):
                messagebox.showerror("Error", "Selected file does not exist")
                return
            
            import_format = format_var.get()
            merge_mode = merge_var.get()
            skip_duplicates = skip_duplicates_var.get()
            validate_data = validate_data_var.get()
            
            try:
                if import_format == "zip":
                    self.import_complete_backup(file_path, merge_mode)
                elif import_format == "csv":
                    self.import_csv_data(file_path, merge_mode, skip_duplicates, validate_data)
                elif import_format == "json":
                    self.import_json_data(file_path, merge_mode, skip_duplicates, validate_data)
                
                import_window.destroy()
                self.refresh_student_list()
                messagebox.showinfo("Success", "Data imported successfully!")
                
            except Exception as e:
                messagebox.showerror("Error", f"Import failed: {e}")
        
        ttk.Button(button_frame, text="📥 Import", command=perform_import).pack(side='right', padx=5)
        ttk.Button(button_frame, text="❌ Cancel", command=import_window.destroy).pack(side='right', padx=5)

    def import_complete_backup(self, zip_path, merge_mode):
        """Import complete backup from ZIP file - SIMPLIFIED VERSION"""
        # Create backup of current data if replace mode
        if merge_mode == "replace" and os.path.exists("graduation_data"):
            backup_path = f"graduation_data_backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}.zip"
            try:
                self.export_complete_backup(backup_path)
                messagebox.showinfo("Backup Created", f"Current data backed up to: {backup_path}")
            except Exception as e:
                if not messagebox.askyesno("Warning", 
                    f"Failed to create backup: {e}\n\nContinue with import anyway?"):
                    return
        
        # Create progress window
        progress_window = tk.Toplevel(self.frame)
        progress_window.title("Importing...")
        progress_window.geometry("400x150")
        progress_window.resizable(False, False)
        progress_window.transient(self.frame)
        progress_window.grab_set()
        
        progress_label = ttk.Label(progress_window, text="Extracting backup...")
        progress_label.pack(pady=20)
        
        progress_var = tk.DoubleVar()
        progress_bar = ttk.Progressbar(progress_window, variable=progress_var, maximum=100)
        progress_bar.pack(pady=10, padx=20, fill='x')
        
        progress_window.update()
        
        try:
            # Validate ZIP file first
            with zipfile.ZipFile(zip_path, 'r') as zipf:
                file_list = zipf.namelist()
                
                # Check if this looks like a valid backup
                has_graduation_data = any(f.startswith('graduation_data/') for f in file_list)
                if not has_graduation_data:
                    raise ValueError("Invalid backup file: No graduation_data folder found in ZIP")
                
                progress_label.config(text="Validating backup...")
                progress_var.set(10)
                progress_window.update()
                
                # If replace mode, remove existing graduation_data folder
                if merge_mode == "replace" and os.path.exists("graduation_data"):
                    progress_label.config(text="Removing existing data...")
                    progress_var.set(20)
                    progress_window.update()
                    shutil.rmtree("graduation_data")
                
                # Extract files
                progress_label.config(text="Extracting files...")
                progress_var.set(30)
                progress_window.update()
                
                total_files = len(file_list)
                extracted_files = 0
                
                for file_name in file_list:
                    # Skip metadata file and directories
                    if file_name.endswith('/') or file_name == 'export_metadata.json':
                        continue
                    
                    # Only extract graduation_data files
                    if file_name.startswith('graduation_data/'):
                        target_path = Path(file_name)
                        
                        # For merge mode, skip if file already exists
                        if merge_mode == "merge" and target_path.exists():
                            extracted_files += 1
                            continue
                        
                        # Create parent directories
                        target_path.parent.mkdir(parents=True, exist_ok=True)
                        
                        # Extract file
                        with zipf.open(file_name) as source, open(target_path, 'wb') as target:
                            shutil.copyfileobj(source, target)
                    
                    extracted_files += 1
                    progress = 30 + (extracted_files / total_files) * 60
                    progress_var.set(progress)
                    progress_window.update()
                
                progress_label.config(text="Finalizing import...")
                progress_var.set(95)
                progress_window.update()
                
            # Refresh the current database connection to pick up imported data
            progress_label.config(text="Refreshing database...")
            self.refresh_student_list()  # This will reload data from files
            
            progress_var.set(100)
            progress_label.config(text="Import completed!")
            progress_window.update()
            
            progress_window.after(1500, progress_window.destroy)
            
        except Exception as e:
            progress_window.destroy()
            raise e

    def validate_backup_file(self, zip_path):
        """Validate that the ZIP file is a valid graduation scanner backup"""
        try:
            with zipfile.ZipFile(zip_path, 'r') as zipf:
                file_list = zipf.namelist()
                
                # Check for required structure
                has_graduation_data = any(f.startswith('graduation_data/') for f in file_list)
                has_database_file = any('students_data.json' in f for f in file_list)
                
                return has_graduation_data and has_database_file
        except:
            return False
            
    def import_csv_data(self, csv_path, merge_mode, skip_duplicates, validate_data):
        """Import student data from CSV file"""
        try:
            # Read CSV file
            df = pd.read_csv(csv_path, encoding='utf-8')
            
            # Validate required columns
            required_columns = ['student_id', 'name', 'faculty', 'graduation_level']
            missing_columns = [col for col in required_columns if col not in df.columns]
            
            if missing_columns:
                raise ValueError(f"Missing required columns: {', '.join(missing_columns)}")
            
            # Get existing students
            existing_students = self.database.get_all_students()
            existing_ids = {s.get('student_id') for s in existing_students}
            
            imported_count = 0
            skipped_count = 0
            errors = []
            
            for index, row in df.iterrows():
                try:
                    student_id = str(row['student_id']).strip()
                    
                    # Skip duplicates if option is enabled
                    if skip_duplicates and student_id in existing_ids:
                        skipped_count += 1
                        continue
                    
                    # Validate data if option is enabled
                    if validate_data:
                        if not student_id or not str(row['name']).strip():
                            errors.append(f"Row {index + 1}: Missing student ID or name")
                            continue
                    
                    # Create student data
                    student_data = {
                        "student_id": student_id,
                        "name": str(row['name']).strip(),
                        "faculty": str(row['faculty']).strip(),
                        "graduation_level": str(row['graduation_level']).strip(),
                        "registered_time": datetime.now().isoformat(),
                        "photo_path": "",
                        "qr_code_path": "",
                        "face_encoding": []
                    }
                    
                    # Add optional columns if they exist
                    if 'photo_path' in df.columns and pd.notna(row['photo_path']):
                        student_data['photo_path'] = str(row['photo_path'])
                    
                    if 'qr_code_path' in df.columns and pd.notna(row['qr_code_path']):
                        student_data['qr_code_path'] = str(row['qr_code_path'])
                    
                    if 'registration_date' in df.columns and pd.notna(row['registration_date']):
                        student_data['registered_time'] = str(row['registration_date'])
                    
                    # Save student
                    self.database.add_student(student_data)
                    imported_count += 1
                    existing_ids.add(student_id)
                    
                except Exception as e:
                    errors.append(f"Row {index + 1}: {str(e)}")
            
            # Show import summary
            summary = f"Import completed!\n\nImported: {imported_count} students"
            if skipped_count > 0:
                summary += f"\nSkipped (duplicates): {skipped_count}"
            if errors:
                summary += f"\nErrors: {len(errors)}"
                if len(errors) <= 5:
                    summary += "\n\nErrors:\n" + "\n".join(errors)
                else:
                    summary += f"\n\nFirst 5 errors:\n" + "\n".join(errors[:5])
            
            messagebox.showinfo("Import Summary", summary)
            
        except Exception as e:
            raise Exception(f"CSV import error: {str(e)}")

    def import_json_data(self, json_path, merge_mode, skip_duplicates, validate_data):
        """Import student data from JSON file"""
        try:
            with open(json_path, 'r', encoding='utf-8') as f:
                data = json.load(f)
            
            # Handle different JSON structures
            if isinstance(data, dict):
                if 'students' in data:
                    students_data = data['students']
                elif 'data' in data:
                    students_data = data['data']
                else:
                    students_data = [data]  # Single student
            elif isinstance(data, list):
                students_data = data
            else:
                raise ValueError("Invalid JSON format")
            
            # Get existing students
            existing_students = self.database.get_all_students()
            existing_ids = {s.get('student_id') for s in existing_students}
            
            imported_count = 0
            skipped_count = 0
            errors = []
            
            for i, student_data in enumerate(students_data):
                try:
                    if not isinstance(student_data, dict):
                        errors.append(f"Student {i + 1}: Invalid data format")
                        continue
                    
                    student_id = str(student_data.get('student_id', '')).strip()
                    
                    # Skip duplicates if option is enabled
                    if skip_duplicates and student_id in existing_ids:
                        skipped_count += 1
                        continue
                    
                    # Validate required fields
                    required_fields = ['student_id', 'name', 'faculty', 'graduation_level']
                    if validate_data:
                        missing_fields = [field for field in required_fields 
                                        if not student_data.get(field, '').strip()]
                        if missing_fields:
                            errors.append(f"Student {i + 1}: Missing fields: {', '.join(missing_fields)}")
                            continue
                    
                    # Ensure all required fields exist
                    processed_student = {
                        "student_id": student_id,
                        "name": str(student_data.get('name', '')).strip(),
                        "faculty": str(student_data.get('faculty', '')).strip(),
                        "graduation_level": str(student_data.get('graduation_level', '')).strip(),
                        "registered_time": student_data.get('registered_time') or student_data.get('registration_date') or datetime.now().isoformat(),
                        "photo_path": student_data.get('photo_path', ''),
                        "qr_code_path": student_data.get('qr_code_path', ''),
                        "face_encoding": student_data.get('face_encoding', [])
                    }
                    
                    # Save student
                    self.database.add_student(processed_student)
                    imported_count += 1
                    existing_ids.add(student_id)
                    
                except Exception as e:
                    errors.append(f"Student {i + 1}: {str(e)}")
            
            # Show import summary
            summary = f"Import completed!\n\nImported: {imported_count} students"
            if skipped_count > 0:
                summary += f"\nSkipped (duplicates): {skipped_count}"
            if errors:
                summary += f"\nErrors: {len(errors)}"
                if len(errors) <= 5:
                    summary += "\n\nErrors:\n" + "\n".join(errors)
                else:
                    summary += f"\n\nFirst 5 errors:\n" + "\n".join(errors[:5])
            
            messagebox.showinfo("Import Summary", summary)
            
        except Exception as e:
            raise Exception(f"JSON import error: {str(e)}")
        